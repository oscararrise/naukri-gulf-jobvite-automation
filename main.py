import os
import json
import copy
import traceback
import requests

from datetime import datetime
from openpyxl import Workbook, load_workbook

from statics import *


LOG_FILE_NAME = "naukri_log.xlsx"


def log_job_result(payload, result, filename=LOG_FILE_NAME):
    exists = os.path.exists(filename)

    if exists:
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    if "Success" not in wb.sheetnames:
        ws_s = wb.active
        ws_s.title = "Success"
        ws_e = wb.create_sheet("Error")
    else:
        ws_s = wb["Success"]
        ws_e = wb["Error"] if "Error" in wb.sheetnames else wb.create_sheet("Error")

    status = (
        "success"
        if isinstance(result, dict) and result.get("status") == "success"
        else "error"
    )

    sheet = ws_s if status == "success" else ws_e

    timestamp = datetime.utcnow().isoformat()
    job_title = (
        payload
        .get("JobPositionPosting", {})
        .get("JobPositionInformation", {})
        .get("JobPositionTitle", "")
    )
    job_id = (
        payload
        .get("JobPositionPosting", {})
        .get("JobPositionPostingID", "")
    )

    sheet.append([
        timestamp,
        job_title,
        job_id,
        status,
        json.dumps(result, ensure_ascii=False),
    ])

    wb.save(filename)


def should_post_to_job_boards(custom_fields):
    for custom_field in custom_fields or []:
        if custom_field.get("fieldCode") == "post_to_job_boards":
            return True
    return False


def get_location_type(custom_fields):
    location_value = ""

    if custom_fields:
        location_value = custom_fields[0].get("value", "") or ""

    return (
        location_value
        .replace("On site", "ONSITE")
        .replace("Remote", "REMOTE")
        .replace("Hybrid", "HYBRID")
    )


def build_naukri_payload(requisition):
    job_title = requisition.get("title", "") or ""
    job_id = requisition.get("requisitionId", "") or ""
    custom_fields = requisition.get("customField", []) or []

    payload = copy.deepcopy(PAYLOAD_TEMPLATE)

    payload["JobPositionPosting"]["JobPositionInformation"]["JobPositionTitle"] = job_title
    payload["JobPositionPosting"]["JobPositionPostingID"] = job_id

    location_type = get_location_type(custom_fields)

    payload["JobPositionPosting"]["JobPositionInformation"]["JobPositionDescription"]["JobPositionLocation"]["LocationType"] = location_type

    summary = requisition.get("description", "") or ""

    payload["JobPositionPosting"]["JobPositionInformation"]["JobPositionDescription"]["SummaryText"] = summary
    payload["JobPositionPosting"]["JobPositionInformation"]["JobPositionRequirements"]["SummaryText"] = summary

    apply_link = requisition.get("applyLink", "") or ""

    payload["JobPositionPosting"]["HowToApply"]["ApplicationMethods"]["ByWeb"]["URL"] = apply_link

    return payload, job_title, job_id


def publish_job_naukrigulf():
    try:
        response = requests.get(
            "https://api.jobvite.com/api/v2/job",
            headers=headers,
            params=params,
            timeout=30,
        )
        response.raise_for_status()

        data = response.json()
        requisitions = data.get("requisitions", [])

        print(f"Total requisitions retrieved: {len(requisitions)}")

        for index, requisition in enumerate(requisitions, start=1):
            job_title = requisition.get("title", "")
            posting_type = requisition.get("postingType", "")
            custom_fields = requisition.get("customField", []) or []

            print(f"Processing job {index}/{len(requisitions)} - {job_title}")

            if posting_type != "External":
                print(f"[SKIP] Job is not External: {job_title}")
                continue

            if not should_post_to_job_boards(custom_fields):
                print(f"[SKIP] post_to_job_boards field not found: {job_title}")
                continue

            payload, var_job_title, var_job_id = build_naukri_payload(requisition)

            if not var_job_title or not var_job_id:
                print(f"[SKIP] Missing job title or requisition ID: {job_title}")
                continue

            if len(var_job_title) > 70:
                print(f"[SKIP] Job title exceeds 70 characters: {var_job_title}")
                continue

            signature, en_access_key = create_keys(var_job_title, var_job_id)

            result = send_post_API(payload, signature, en_access_key)
            log_job_result(payload, result)

            print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print(result)
            print(f"Job {var_job_title} published on Naukri Gulf.")
            print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

    except Exception as e:
        print("Error naukri post job:", str(e))
        print(traceback.format_exc())


if __name__ == "__main__":
    publish_job_naukrigulf()